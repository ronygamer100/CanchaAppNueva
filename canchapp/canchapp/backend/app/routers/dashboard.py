from datetime import date, datetime, timedelta, time
from app.core.timezone import today_peru, now_peru
from typing import List, Optional, Literal
from collections import defaultdict
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core.database import get_db
from app.core.deps import get_current_owner
from app.models.owner import Owner
from app.models.venue import Venue
from app.models.court import Court
from app.models.reservation import Reservation, ReservationStatus
from app.services.auto_confirm import sweep_auto_confirms


router = APIRouter(prefix="/api", tags=["dashboard"])


# ============ Schemas internos ============
class WeekReservationItem(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    court_id: int
    court_nombre: str
    fecha: date
    hora_inicio: time
    hora_fin: time
    jugador_nombre: str
    jugador_whatsapp: str
    estado: ReservationStatus
    horas: int
    created_at: datetime


class WeekReservationsOut(BaseModel):
    week_start: date
    week_end: date
    items: List[WeekReservationItem]


class KPIs(BaseModel):
    ingresos_mes: float
    reservas_confirmadas: int
    reservas_canceladas: int
    reservas_rechazadas: int
    horario_popular: Optional[str] = None
    cancha_top: Optional[str] = None
    tasa_ocupacion_pct: float


class DailyIngreso(BaseModel):
    fecha: date
    monto: float


class HourlyOcupacion(BaseModel):
    hora: int
    reservas: int


class MetricsOut(BaseModel):
    period: str  # "month" | "week"
    range_start: date
    range_end: date
    kpis: KPIs
    ingresos_diarios: List[DailyIngreso]
    ocupacion_horaria: List[HourlyOcupacion]


# ============ Helpers ============
def _venue_or_404(db: Session, venue_id: int, owner: Owner) -> Venue:
    venue = db.query(Venue).filter(
        Venue.id == venue_id, Venue.owner_id == owner.id
    ).first()
    if not venue:
        raise HTTPException(status_code=404, detail="Negocio no encontrado")
    return venue


def _horas_reserva(r: Reservation) -> int:
    inicio = datetime.combine(r.fecha, r.hora_inicio)
    fin = datetime.combine(r.fecha, r.hora_fin)
    return int((fin - inicio).total_seconds() // 3600)


def _monday_of(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _compute_range(
    period: str, year: Optional[int], month: Optional[int], week_start: Optional[date],
) -> tuple[date, date]:
    """Devuelve (first_day, last_day) según el período."""
    if period == "month":
        if year is None or month is None:
            today = today_peru()
            year, month = today.year, today.month
        first = date(year, month, 1)
        if month == 12:
            nf = date(year + 1, 1, 1)
        else:
            nf = date(year, month + 1, 1)
        last = nf - timedelta(days=1)
        return first, last
    else:  # week
        start = _monday_of(week_start or today_peru())
        return start, start + timedelta(days=6)


def _build_metrics(
    db: Session, venue: Venue, first_day: date, last_day: date,
) -> tuple[KPIs, List[DailyIngreso], List[HourlyOcupacion], list]:
    """Calcula KPIs + series. Devuelve también las filas confirmadas para reusar."""
    total_dias = (last_day - first_day).days + 1
    canchas = db.query(Court).filter(Court.venue_id == venue.id).all()
    cancha_by_id = {c.id: c for c in canchas}

    # Confirmadas
    rows = (
        db.query(Reservation, Court)
        .join(Court, Reservation.court_id == Court.id)
        .filter(
            Court.venue_id == venue.id,
            Reservation.fecha >= first_day,
            Reservation.fecha <= last_day,
            Reservation.estado == ReservationStatus.CONFIRMADA,
        )
        .all()
    )

    counts_by_status = dict(
        db.query(Reservation.estado, func.count(Reservation.id))
        .join(Court, Reservation.court_id == Court.id)
        .filter(
            Court.venue_id == venue.id,
            Reservation.fecha >= first_day,
            Reservation.fecha <= last_day,
        )
        .group_by(Reservation.estado)
        .all()
    )

    confirmadas = counts_by_status.get(ReservationStatus.CONFIRMADA, 0)
    canceladas = counts_by_status.get(ReservationStatus.CANCELADA, 0)
    rechazadas = counts_by_status.get(ReservationStatus.RECHAZADA, 0)

    ingresos_por_dia: dict[date, float] = defaultdict(float)
    ingresos_por_cancha: dict[int, float] = defaultdict(float)
    ocupacion_por_hora: dict[int, int] = defaultdict(int)
    total_ingresos = 0.0
    total_horas_reservadas = 0

    for r, court in rows:
        horas = _horas_reserva(r)
        monto = court.precio_hora * horas
        ingresos_por_dia[r.fecha] += monto
        ingresos_por_cancha[court.id] += monto
        total_ingresos += monto
        total_horas_reservadas += horas
        for h in range(r.hora_inicio.hour, r.hora_fin.hour):
            ocupacion_por_hora[h] += 1

    horario_popular = None
    if ocupacion_por_hora:
        h = max(ocupacion_por_hora, key=ocupacion_por_hora.get)
        horario_popular = f"{h:02d}:00"

    cancha_top = None
    if ingresos_por_cancha:
        top_id = max(ingresos_por_cancha, key=ingresos_por_cancha.get)
        cancha_top = cancha_by_id.get(top_id).nombre if top_id in cancha_by_id else None

    horas_op_dia = max(0, venue.hora_cierre.hour - venue.hora_apertura.hour)
    horas_disp = horas_op_dia * total_dias * len(canchas)
    tasa = (total_horas_reservadas / horas_disp * 100) if horas_disp else 0

    ingresos_diarios = []
    cur = first_day
    while cur <= last_day:
        ingresos_diarios.append(DailyIngreso(fecha=cur, monto=round(ingresos_por_dia[cur], 2)))
        cur = cur + timedelta(days=1)

    ocupacion_horaria = []
    for h in range(venue.hora_apertura.hour, venue.hora_cierre.hour):
        ocupacion_horaria.append(HourlyOcupacion(hora=h, reservas=ocupacion_por_hora[h]))

    kpis = KPIs(
        ingresos_mes=round(total_ingresos, 2),
        reservas_confirmadas=confirmadas,
        reservas_canceladas=canceladas,
        reservas_rechazadas=rechazadas,
        horario_popular=horario_popular,
        cancha_top=cancha_top,
        tasa_ocupacion_pct=round(tasa, 1),
    )
    return kpis, ingresos_diarios, ocupacion_horaria, rows


# ============ Endpoints ============
@router.get("/venues/{venue_id}/week-reservations", response_model=WeekReservationsOut)
def week_reservations(
    venue_id: int,
    week_start: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    owner: Owner = Depends(get_current_owner),
):
    venue = _venue_or_404(db, venue_id, owner)
    sweep_auto_confirms(db, owner_id=owner.id)

    start = _monday_of(week_start or today_peru())
    end = start + timedelta(days=6)

    rows = (
        db.query(Reservation, Court.nombre.label("court_nombre"))
        .join(Court, Reservation.court_id == Court.id)
        .filter(
            Court.venue_id == venue.id,
            Reservation.fecha >= start,
            Reservation.fecha <= end,
            Reservation.estado.in_([
                ReservationStatus.PENDIENTE,
                ReservationStatus.CONFIRMADA,
            ]),
        )
        .order_by(Reservation.fecha, Reservation.hora_inicio)
        .all()
    )

    items = [
        WeekReservationItem(
            id=r.id, court_id=r.court_id, court_nombre=court_nombre,
            fecha=r.fecha, hora_inicio=r.hora_inicio, hora_fin=r.hora_fin,
            jugador_nombre=r.jugador_nombre, jugador_whatsapp=r.jugador_whatsapp,
            estado=r.estado, horas=_horas_reserva(r),
            created_at=r.created_at,
        )
        for r, court_nombre in rows
    ]

    return WeekReservationsOut(week_start=start, week_end=end, items=items)


@router.get("/venues/{venue_id}/metrics", response_model=MetricsOut)
def venue_metrics(
    venue_id: int,
    period: Literal["month", "week"] = Query("month"),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None, ge=1, le=12),
    week_start: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    owner: Owner = Depends(get_current_owner),
):
    venue = _venue_or_404(db, venue_id, owner)
    sweep_auto_confirms(db, owner_id=owner.id)

    first_day, last_day = _compute_range(period, year, month, week_start)
    kpis, ingresos_diarios, ocupacion_horaria, _ = _build_metrics(db, venue, first_day, last_day)

    return MetricsOut(
        period=period,
        range_start=first_day,
        range_end=last_day,
        kpis=kpis,
        ingresos_diarios=ingresos_diarios,
        ocupacion_horaria=ocupacion_horaria,
    )


@router.get("/venues/{venue_id}/metrics/export")
def export_metrics_xlsx(
    venue_id: int,
    period: Literal["month", "week"] = Query("month"),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None, ge=1, le=12),
    week_start: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    owner: Owner = Depends(get_current_owner),
):
    """Devuelve un .xlsx con dos hojas: KPIs y Detalle de reservas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    venue = _venue_or_404(db, venue_id, owner)
    sweep_auto_confirms(db, owner_id=owner.id)

    first_day, last_day = _compute_range(period, year, month, week_start)
    kpis, ingresos_diarios, _, confirmadas_rows = _build_metrics(db, venue, first_day, last_day)

    # Traer TODAS las reservas del rango (todos los estados) para el detalle
    todas = (
        db.query(Reservation, Court)
        .join(Court, Reservation.court_id == Court.id)
        .filter(
            Court.venue_id == venue.id,
            Reservation.fecha >= first_day,
            Reservation.fecha <= last_day,
        )
        .order_by(Reservation.fecha, Reservation.hora_inicio)
        .all()
    )

    wb = Workbook()

    # -------- Hoja 1: KPIs --------
    ws1 = wb.active
    ws1.title = "Resumen"

    # Estilos
    header_font = Font(bold=True, color="FFFFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="FF0E3B2E")
    title_font = Font(bold=True, size=16)
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="FFCCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws1["A1"] = venue.nombre
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:D1")

    periodo_label = "Mes" if period == "month" else "Semana"
    ws1["A2"] = f"{periodo_label}: {first_day.strftime('%d/%m/%Y')} — {last_day.strftime('%d/%m/%Y')}"
    ws1.merge_cells("A2:D2")

    ws1["A4"] = "Indicador"
    ws1["B4"] = "Valor"
    for cell in (ws1["A4"], ws1["B4"]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    kpi_rows = [
        ("Ingresos del período", f"S/ {kpis.ingresos_mes:.2f}"),
        ("Reservas confirmadas", kpis.reservas_confirmadas),
        ("Reservas canceladas", kpis.reservas_canceladas),
        ("Reservas rechazadas", kpis.reservas_rechazadas),
        ("Hora más popular", kpis.horario_popular or "—"),
        ("Cancha más rentable", kpis.cancha_top or "—"),
        ("Tasa de ocupación", f"{kpis.tasa_ocupacion_pct}%"),
    ]
    for i, (label, value) in enumerate(kpi_rows, start=5):
        ws1[f"A{i}"] = label
        ws1[f"A{i}"].font = bold
        ws1[f"B{i}"] = value
        for col in ("A", "B"):
            ws1[f"{col}{i}"].border = border

    # Tabla de ingresos diarios
    base_row = 5 + len(kpi_rows) + 2
    ws1[f"A{base_row}"] = "Ingresos por día"
    ws1[f"A{base_row}"].font = title_font
    ws1.merge_cells(f"A{base_row}:B{base_row}")

    ws1[f"A{base_row + 1}"] = "Fecha"
    ws1[f"B{base_row + 1}"] = "Monto (S/)"
    for c in (f"A{base_row + 1}", f"B{base_row + 1}"):
        ws1[c].font = header_font
        ws1[c].fill = header_fill

    for i, d in enumerate(ingresos_diarios):
        r = base_row + 2 + i
        ws1[f"A{r}"] = d.fecha.strftime("%d/%m/%Y")
        ws1[f"B{r}"] = d.monto
        ws1[f"B{r}"].number_format = '"S/ "#,##0.00'
        ws1[f"A{r}"].border = border
        ws1[f"B{r}"].border = border

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 22

    # -------- Hoja 2: Detalle de reservas --------
    ws2 = wb.create_sheet("Reservas")

    headers = [
        "Fecha", "Hora inicio", "Hora fin", "Horas",
        "Cancha", "Jugador", "WhatsApp", "Estado",
        "Monto cancha (S/)", "Adelanto (S/)", "Creada el",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for i, (r, court) in enumerate(todas, start=2):
        horas = _horas_reserva(r)
        monto = court.precio_hora * horas
        adelanto = court.adelanto_monto * horas
        ws2.cell(row=i, column=1, value=r.fecha.strftime("%d/%m/%Y"))
        ws2.cell(row=i, column=2, value=r.hora_inicio.strftime("%H:%M"))
        ws2.cell(row=i, column=3, value=r.hora_fin.strftime("%H:%M"))
        ws2.cell(row=i, column=4, value=horas)
        ws2.cell(row=i, column=5, value=court.nombre)
        ws2.cell(row=i, column=6, value=r.jugador_nombre)
        ws2.cell(row=i, column=7, value=r.jugador_whatsapp)
        ws2.cell(row=i, column=8, value=r.estado.value)
        ws2.cell(row=i, column=9, value=monto).number_format = '"S/ "#,##0.00'
        ws2.cell(row=i, column=10, value=adelanto).number_format = '"S/ "#,##0.00'
        ws2.cell(row=i, column=11, value=r.created_at.strftime("%d/%m/%Y %H:%M"))

        # Color de fila según estado
        if r.estado == ReservationStatus.CONFIRMADA:
            fill = PatternFill("solid", fgColor="FFDAF4DF")
        elif r.estado == ReservationStatus.PENDIENTE:
            fill = PatternFill("solid", fgColor="FFFFF8E1")
        elif r.estado == ReservationStatus.RECHAZADA:
            fill = PatternFill("solid", fgColor="FFFEE4E2")
        else:  # cancelada
            fill = PatternFill("solid", fgColor="FFEEEEEE")
        for col in range(1, 12):
            ws2.cell(row=i, column=col).fill = fill

    # Anchos
    widths = [12, 10, 10, 6, 16, 22, 16, 12, 16, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    # Total al final
    if todas:
        total_row = len(todas) + 2
        ws2.cell(row=total_row + 1, column=8, value="TOTAL CONFIRMADAS").font = bold
        total_confirmadas = sum(
            court.precio_hora * _horas_reserva(r)
            for r, court in todas if r.estado == ReservationStatus.CONFIRMADA
        )
        ws2.cell(row=total_row + 1, column=9, value=total_confirmadas).number_format = '"S/ "#,##0.00'
        ws2.cell(row=total_row + 1, column=9).font = bold

    # Serializar
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    label = f"{first_day.strftime('%Y%m%d')}-{last_day.strftime('%Y%m%d')}"
    safe_slug = "".join(c for c in venue.slug if c.isalnum() or c in "-_")
    filename = f"{safe_slug}-{label}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
